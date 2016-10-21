# -*- coding: utf-8 -*-
"""
Created on Mon Oct 17 16:29:55 2016

@author: guillaume
"""

from openpyxl import Workbook


def comp2excel(comp_summary,deps):
    """

    Parameters
    -----------
    
    comp_summary:

            keys: 'name', 'build', 'dep'
            
            comp_summary[0]['name']  = 'Brain morphological features'
            comp_summary[0]['build'] = 'Data Factory (DF)'
            comp_summary[0]['dep']   = ['Lille Hospital', 'Tel Aviv Hospital', 'Milano Hospital', 'Freiburg Hospital', 'CHUV Hospital', 'Data Storage']
            
            
    deps:   set of dependencies: ['Image & Genetic Viewer', 'Predictive Models', 'Common Variables & Metadata',.... ]

    """    
    
    wb  = Workbook()
    ws1 = wb.create_sheet("summary") #


    # Add dependency names in columne
    r = 2    
    for dep in deps:
        ws1.cell(row=r, column=1, value=dep)    
        r = r + 1


    # Add dependency names in columne
    col = 2    
    for summ in comp_summary:
        ws1.cell(row=1, column=col, value=summ['name'])    
        
        idx = get_dep_indicies(summ,deps)
        
        print idx
        
        if idx:
            for i in idx:
                ws1.cell(row=i+2, column=col, value=['1'])    
        
        col = col + 1
        


    return wb
    
        
def get_dep_indicies(summ,deps):
    """
        For a component name find the row indicies of the dependencies.
    """
    
    dep      = summ['dep']
    indicies = []
 
    for d in dep:

        for i in range(0,len(deps)):
            if d == deps[0]:
                indicies.extend(i)
            
    return indicies
        
    
    
    