from openpyxl import load_workbook

def import_excel(filename):
    """
        Imports an excel sheet
    
    """
    
    wb       = load_workbook(filename, use_iterators=True)
    sh = wb["data"]

    return sh
    
    
    
def get_col_indicies(sh,col_names):
    """
        given a list of column names (first row) return all column indicies 
        which mache the names.    
    """
    
    indices = dict()
    
    # Get numer of columns
    num_cols = sh.max_column
    
   
    #Iterate through worksheet and print cell contents
    for j in range(1,num_cols):
        
        cell_name = str(sh.cell(row=1,column=j).value)
        
        if cell_name in col_names:
            
            if cell_name in indices:
                # append the new number to the existing array at this slot
                indices[cell_name].append(j)
            else:
                # create a new array in this slot
                indices[cell_name] = [j]
            
    return indices            
    
    
def col_not_checked(sh):
    """
        Check the entries of a column containes checked or unchecked    

    """
    token = ['Unchecked','Checked']
    indicies = []
    
    # Get numer of columns
    num_cols = sh.max_column
    
    for j in range(1,num_cols):
       cell_name = str(sh.cell(row=2,column=j).value)
       
       if cell_name not in token:
           indicies.append(j)
           
    return indicies       


def extrat_data_row(sh,row_id,col_index):

        data_val  = []
        col_names = []
               
        
        for j in col_index:
            
            col_name = str(sh.cell(row=1,column=j).value)
            val      = sh.cell(row=row_id,column=j).value

            if isinstance(val,unicode):
                data_j = val.encode('ascii','ignore')
            else:
                data_j = str(val)
                
                
            #val = ws.cell(row=8,column=index[4]).value
            
            col_names.append(col_name)
            data_val.append(data_j)
            
            
        return data_val,col_names        



















    
    
    
        
        
        
    



    