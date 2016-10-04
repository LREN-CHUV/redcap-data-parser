from openpyxl import load_workbook
import re



def import_excel(filename):
    """
        Imports an excel sheet
    
    """
    
    wb = load_workbook(filename)
    sh = wb["data"]

    return sh
    
    
    
def get_col_indicies(sh,col_names):
    """
        given a list of column names (first row) return all column indicies 
        which mache the names.    
    """
    
    indices = dict()
    
    # Get numer of columns
    num_cols = sh.max_column
    
   
    #Iterate through worksheet and print cell contents
    for j in range(1,num_cols):
        
        cell_name = str(sh.cell(row=1,column=j).value)
        
        if cell_name in col_names:
            
            if cell_name in indices:
                # append the new number to the existing array at this slot
                indices[cell_name].append(j)
            else:
                # create a new array in this slot
                indices[cell_name] = [j]
            
    return indices            
    
    
def col_not_checked(sh):
    """
        Check the entries of a column containes checked or unchecked    

    """
    token = ['Unchecked','Checked']
    indicies = []
    
    # Get numer of columns
    num_cols = sh.max_column
    
    for j in range(1,num_cols):
       cell_name = str(sh.cell(row=2,column=j).value)
       
       if cell_name not in token:
           indicies.append(j)
           
    return indicies       


def get_question_type(str1):
    if len(re.findall('\\b(select one or many)\\b', str1)) > 0:
        return '(select one or many)'
    elif len(re.findall('\\b(select one)\\b', str1)) > 0:
        return '(select one)' 
    else:
        return "'(choice='"

def get_question(str1,str2):
    index = str1.find(str2)
    if index > 1:
        return str1[0:index-1]
    else:
        return ''   

def get_answer(str1,str2='choice='):
    len_str1 = len(str1)
    len_str2 = len(str2)
    index = str1.find(str2) + len_str2
    
    if index > 1 & index < len_str1:
        return str1[index+1:len_str1-2]
    else:
        return ''   


def extrat_row_no_check(sh,row_id,col_index):
   """
   
   """
   data_val  = []
   col_names = []
   
   for j in col_index:
       col_name = str(sh.cell(row=1,column=j).value)
       val      =     sh.cell(row=row_id,column=j).value
       if isinstance(val,unicode):
           data_j = val.encode('ascii','ignore')
       else:
           data_j = str(val)
   
       col_names.append(col_name)
       data_val.append(data_j)
            
            
   return data_val,col_names        




def extrat_row(ws,row_id):
   """
   
   """
   col_names = []
   data_val  = []

   num_col = ws.max_column

   for j in range(1,num_col):
       col_name = str(ws.cell(row=1,column=j).value)
       val      =     ws.cell(row=row_id,column=j).value

       # ASCII check
       if isinstance(val,unicode):
           data_j = val.encode('ascii','ignore')
       else:
           data_j = str(val)
           
           
       append   = True
    
       if data_j == 'Checked':
           question_type = get_question_type(col_name)        
           question = get_question(col_name,question_type)    
           
           answer   = get_answer(col_name)


           col_name = question
           data_j   = answer     
           
           append   = True
       elif data_j == 'Unchecked':
           append   = False
       else:
           append   = True

       if col_name == 'To which building block your component belongs to ? (select one)':
           col_name = 'To which building block your component belongs to ?'         
        

       if append:  
           col_names.append(col_name)
           data_val.append(data_j) 
       
   return col_names,data_val    














    
    
    
        
        
        
    



    